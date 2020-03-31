import requests
import time
import random
import datetime
import csv
import os
from bs4 import BeautifulSoup
from requests import HTTPError
from lxml import html
from openpyxl import workbook, load_workbook
from .models import RusprofileParser, RusprofileParserData, RusprofileSettings
from contacts_parser.models import PageData
from multiprocessing import Pool
from openpyxl import Workbook
from openpyxl.styles import NamedStyle, Font, Border, Side, Alignment, colors
from openpyxl.utils.exceptions import IllegalCharacterError
import concurrent.futures
from site_engine.settings import BASE_DIR


def get_headers():
    with open(os.path.join(BASE_DIR, 'user-agents/user_agents_for_chrome_pk.txt')) as u_a:
        user_agents = [row.strip() for row in u_a.readlines()]

    headers = {
        'accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.9',
        'accept-encoding': 'gzip, deflate',
        'accept-language': 'ru,ru-RU;q=0.9,kk-RU;q=0.8,kk;q=0.7,uz-RU;q=0.6,uz;q=0.5,en-RU;q=0.4,en-US;q=0.3,en;q=0.2',
        'cache-control': 'max-age=0',
        'connection': 'keep-alive',
        'sec-fetch-dest': 'document',
        'sec-fetch-mode': 'navigate',
        'sec-fetch-site': 'none',
        'sec-fetch-user': '?1',
        'upgrade-insecure-requests': '1',
        'user-Agent': random.choice(user_agents),
    }
    return headers


def check_proxy(proxy, headers):
    try:
        request = requests.get(
            url='https://check-host.net/ip',
            headers=headers,
            proxies=proxy
        )
        print(request.text)

    except HTTPError as http_err:
        print(f'HTTP error : {http_err}')
    except Exception as err:
        print(f'Other err : {err}')
    except TimeoutError as time_err:
        print(f'Timeout error : {time_err}')
    except:
        print('Broken')


def get_data(key, proxy):
    status = False
    json_data = {}
    settings = RusprofileSettings.objects.all()[0]

    try:

        headers = get_headers()

        request = requests.get(
            url=f'https://www.rusprofile.ru/ajax.php?&query={key}&action=search',
            stream=True,
            headers=headers,
            proxies=proxy.get_new_proxy(),
            timeout=settings.waiting_time
        )

        json_data = request.json()
        status = True
    except HTTPError as http_err:
        # print(f'HTTP error : {http_err}')
        pass
    except Exception as err:
        # print(f'Other err : {err}')
        pass
    except TimeoutError as time_err:
        # print(f'Timeout error : {time_err}')
        pass
    except:
        # print('Broken')
        pass

    if status is True:
        if json_data['message'] == 'Подтвердите, что вы не робот':
            print('РОБОТ!!!')
            time.sleep(6)
            get_data(key, proxy)
        else:
            try:
                if int(json_data['ul_count']) != 0:
                    s_type = 'ul'

                elif int(json_data['ip_count']) != 0:
                    s_type = 'ip'

                else:
                    s_type = False

                if s_type is not False:

                    for rez in json_data[s_type]:
                        link = rez['link']

                        soup = get_soup(f'https://www.rusprofile.ru{link}', settings, proxy)
                        company_data = get_data_from_soup(soup)

                        return company_data

            except KeyError:
                # print(json_data)
                return False
    else:
        return False


def get_soup(page_url, settings, proxy):
    status = True
    count = 1
    while count < 5 or status is False:
        # print(f'URL {page_url} | Проход : {count}')
        time.sleep(random.uniform(2, 4))
        try:

            response = requests.get(
                page_url,
                headers=get_headers(),
                proxies=proxy.get_new_proxy(),
                stream=True,
                timeout=settings.waiting_time
            )

            rez_soup = BeautifulSoup(response.text, 'lxml')
            check_for_robots = False
            if check_for_robots:
                print('Возможно запрос был заблокирован проверкой на робота')
                status = False
                return False
            else:
                status = True
                return rez_soup, response.text

        except HTTPError as http_err:
            # print(f'HTTP error : {http_err}')
            pass
        except Exception as err:
            # print(f'Other err : {err}')
            pass
        except TimeoutError as time_err:
            # print(f'Timeout error : {time_err}')
            pass
        except:
            # print('Broken')
            pass
        finally:
            count += 1
    if count > 5:
        return False


def get_data_from_soup(d):

    page_soup = d[0]
    page_text = d[1]
    tree = html.fromstring(page_text)

    try:
        company_name = page_soup.select_one("*.company-name").getText(strip=True)
    except AttributeError:
        company_name = ''

    try:
        st = page_soup.select_one("*.company-status").getText(strip=True)
    except AttributeError:
        st = ''

    try:
        ogrn = page_soup.select_one("*#clip_ogrn").getText(strip=True)
    except AttributeError:
        ogrn = ''

    try:
        inn = page_soup.select_one("*#clip_inn").getText(strip=True)
    except AttributeError:
        inn = ''

    try:
        kpp = page_soup.select_one("*#clip_kpp").getText(strip=True)
    except AttributeError:
        kpp = ''

    try:
        reg_date = tree.xpath(
            "//dt[contains(text(),'Дата регистрации')]/following::dd[1]"
        )[0].text.replace('\n', '').replace('  ', '')
    except:
        reg_date = ''

    try:
        address = page_soup.select_one("address.company-info__text").getText(strip=True)
    except AttributeError:
        try:
            address = tree.xpath(
                "//dt[@class='company-info__title' and contains(text(),'Адрес')]/following::dd[1]"
            )[0].text_content().replace('\n', '').replace('  ', '')
        except (AttributeError, IndexError):
            address = ''

    try:
        kap = tree.xpath(
            "//dt[contains(text(),'Уставный капитал')]/following::dd[1]/span[1]"
        )[0].text.replace('\n', '').replace('  ', '')
    except (AttributeError, IndexError):
        kap = ''

    try:
        supervisor = page_soup.select_one("div.hidden-parent span.company-info__text").getText(strip=True)
    except AttributeError:
        supervisor = ''

    try:
        workers = tree.xpath(
            "//dt[contains(text(),'Среднесписочная численность ')]/following::dd[1]"
        )[0].text.replace('\n', '').replace('  ', '')
    except (AttributeError, IndexError):
        workers = ''

    try:
        nalog_rezim = tree.xpath("//dt[contains(text(),'налоговый режим ')]/following::dd[1]")[0].text.replace('\n', '').replace('  ', '')
    except (AttributeError, IndexError):
        nalog_rezim = ''

    try:
        company_type = str(tree.xpath(
            "//span[contains(text(),'Основной вид деятельности')]/following::span[1]"
        )[0].text).strip().replace('  ', '')
    except (AttributeError, IndexError):
        try:
            company_type = str(tree.xpath(
                "//div[@class='tile-item__text-title' and contains(text(),'Основной')]/following::p[1]"
            )[0].text).strip().replace('  ', '')
        except (AttributeError, IndexError):
            company_type = ''

    try:
        status = tree.xpath("//span[contains(text(),'Статус:')]")[0].text
    except (AttributeError, IndexError):
        try:
            status = tree.xpath("//dt[contains(text(),'Категория субъекта')]/following::dd[1]")[0].text
        except (AttributeError, IndexError):
            status = ''

    try:
        all_founders = []
        founders = tree.xpath(
            "//a[contains(text(),'Учредители')]/parent::h2/parent::div/div[@class='founder-item']//a/span"
        )
        for founder in founders:
            all_founders.append(str(founder.text).strip().replace('  ', ''))

        rez_founders = ', '.join(all_founders)
    except (AttributeError, IndexError):
        rez_founders = ''

    try:
        revenue = str(tree.xpath(
            "//div[contains(text(), 'Выручка')]/parent::div[1]/following::div[1]"
        )[0].text_content()).replace('\n', '').replace('  ', '')
    except:
        revenue = ''

    try:
        profit = str(tree.xpath(
            "//div[contains(text(), 'Прибыль')]/parent::div[1]/following::div[1]"
        )[0].text_content()).replace('\n', '').replace('  ', '')
    except (AttributeError, IndexError):
        profit = ''

    try:
        nalog = str(tree.xpath(
            "//div[contains(text(), 'Налоги')]/following::div[1]"
        )[0].text_content()).replace('\n', '').replace('  ', '')
    except (AttributeError, IndexError):
        nalog = ''

    results = {
        'name': company_name,
        'st': st,
        'ogrn': ogrn,
        'inn': inn,
        'kpp': kpp,
        'reg_date': reg_date,
        'address': address,
        'kap': kap,
        'supervisor': supervisor,
        'workers': workers,
        'nalog_rezim': nalog_rezim,
        'company_type': company_type,
        'status': status,
        'rez_founders': rez_founders,
        'revenue': revenue,
        'profit': profit,
        'nalog': nalog
    }

    return results


def create_new_data(parser_name, page_data, data):
    try:
        new_data = RusprofileParserData(
            parser_name=parser_name,
            search_data=page_data,
            company_name=data['name'],
            st=data['st'],
            ogrn=data['ogrn'],
            kpp=data['kpp'],
            inn=data['inn'],
            reg_date=data['reg_date'],
            address=data['address'],
            kap=data['kap'],
            supervisor=data['supervisor'],
            workers=data['workers'],
            nalog_rezhim=data['nalog_rezim'],
            company_type=data['company_type'],
            status=data['status'],
            all_founders=data['rez_founders'],
            revenue=data['revenue'],
            profit=data['profit'],
            nalog=data['nalog']
        )
        new_data.save()
    except TypeError:
        pass


def reform_data_from_rusprofile_bd(row):
    row_data = {
        'name': row.company_name,
        'st': row.st,
        'ogrn': row.ogrn,
        'inn': row.inn,
        'kpp': row.kpp,
        'reg_date': row.reg_date,
        'address': row.address,
        'kap': row.kap,
        'supervisor': row.supervisor,
        'workers': row.workers,
        'nalog_rezim': row.nalog_rezhim,
        'company_type': row.company_type,
        'status': row.status,
        'rez_founders': row.all_founders,
        'revenue': row.revenue,
        'profit': row.profit,
        'nalog': row.nalog
    }
    return row_data


def filter_and_search(data):
    if data[1].use_key_words == 'True':
        target = int(data[1].number_of_coincidences)
        key_words = str(data[1].key_words).split(',')
        search_count = 0

        for keys in key_words:
            key = keys.strip().split(' ')
            check = True
            for k in key:
                if check is True:
                    if k.lower() in data[0].title.lower() or k in data[0].description.lower():
                        check = True
                    else:
                        check = False
                        break
            if check is True:
                search_count += 1
    else:
        search_count = 0
        target = 0

    if search_count >= target:
        if len(data[0].inn) > 1 or len(data[0].ogrn) > 1 or len(data[0].kpp) > 1:

            if len(data[0].inn) > 1:
                for i in str(data[0].inn).split(', '):
                    i_search_in_bd = RusprofileParserData.objects.filter(inn__iexact=i)
                    if i_search_in_bd:
                        result = reform_data_from_rusprofile_bd(i_search_in_bd[0])
                    else:
                        result = get_data(i, data[2])
                    if result is not False:
                        create_new_data(data[1], data[0], result)
            elif len(data[0].ogrn) > 1:
                for i in str(data[0].ogrn).split(', '):
                    ogrn_search_in_bd = RusprofileParserData.objects.filter(ogrn__iexact=i)
                    if ogrn_search_in_bd:
                        result = reform_data_from_rusprofile_bd(ogrn_search_in_bd[0])
                    else:
                        result = get_data(i, data[2])
                    if result is not False:
                        create_new_data(data[1], data[0], result)
            elif len(data[0].kpp) > 1:
                for i in str(data[0].kpp).split(', '):
                    kpp_search_in_bd = RusprofileParserData.objects.filter(kpp__iexact=i)
                    if kpp_search_in_bd:
                        result = reform_data_from_rusprofile_bd(kpp_search_in_bd[0])
                    else:
                        result = get_data(i, data[2])
                    if result is not False:
                        create_new_data(data[1], data[0], result)

    data[1].add_done_url()


def write_rusprofile_xlsx_file(search):
    # print('WRITE')
    csv.field_size_limit(100000000)

    file_name = f'{datetime.datetime.today().strftime("%d-%m-%Y_%H-%M")}.xlsx'

    wb = Workbook()

    sheet = wb['Sheet']
    wb.remove(sheet)

    wb.create_sheet('Список сайтов')
    sheet = wb['Список сайтов']

    sheet.column_dimensions['A'].width = 40
    sheet.column_dimensions['B'].width = 30
    sheet.column_dimensions['C'].width = 20
    sheet.column_dimensions['D'].width = 20
    sheet.column_dimensions['E'].width = 20
    sheet.column_dimensions['F'].width = 30
    sheet.column_dimensions['G'].width = 100
    sheet.column_dimensions['H'].width = 30
    sheet.column_dimensions['I'].width = 20
    sheet.column_dimensions['J'].width = 40
    sheet.column_dimensions['K'].width = 20
    sheet.column_dimensions['L'].width = 100
    sheet.column_dimensions['M'].width = 30
    sheet.column_dimensions['N'].width = 50
    sheet.column_dimensions['O'].width = 20
    sheet.column_dimensions['P'].width = 20
    sheet.column_dimensions['Q'].width = 20
    sheet.column_dimensions['R'].width = 30
    sheet.column_dimensions['S'].width = 50
    sheet.column_dimensions['T'].width = 50
    sheet.column_dimensions['U'].width = 50
    sheet.column_dimensions['V'].width = 50
    sheet.column_dimensions['W'].width = 50
    sheet.column_dimensions['X'].width = 200
    sheet.column_dimensions['Y'].width = 200

    search_results = RusprofileParserData.objects.filter(parser_name=search)

    header = NamedStyle(name="header")
    header.font = Font(bold=True, color=colors.RED, size=15)
    header.border = Border(bottom=Side(border_style="thin"))
    header.alignment = Alignment(horizontal="center", vertical="center")
    sheet.append(['Название компании', 'Статус компании', 'ИНН', 'ОГРН', 'КПП', 'Дата регистрации', 'Адрес',
                  'Руководитель', 'Капитал', 'Сотрудники', 'Налоговый режим',
                  'Основной вид деятельности', 'Статус', 'Учредители', 'Выручка', 'Прибыль', 'Налог',
                  'Url', 'Email', 'Телефон', 'ИНН', 'ОГРН', 'КПП', 'Заголовок', 'Описание'])

    header_row = sheet[1]
    for cell in header_row:
        cell.style = header

    r = 0
    for result in search_results:

        data = [result.company_name, result.st, result.inn, result.ogrn, result.kpp, result.reg_date, result.address,
                result.supervisor, result.kap, result.workers, result.nalog_rezhim, result.company_type, result.status,
                result.all_founders, result.revenue, result.profit, result.nalog, result.search_data.page_url,
                result.search_data.emails, result.search_data.phones, result.search_data.inn, result.search_data.ogrn,
                result.search_data.kpp, result.search_data.title.strip(), result.search_data.description.strip()]

        try:
            sheet.append(data)
            r += 1

        except IllegalCharacterError:
            pass

    path = os.path.join(BASE_DIR, f'media/results_files/rusprofile/{search.file_name}')
    if os.path.isdir(path) is False:
        os.makedirs(path, mode=0o777)

    wb.save(f'{path}/{file_name}')

    search.result_file = f'results_files/rusprofile/{search.file_name}/{file_name}'
    search.change_result_file()


class Proxy:
    def __init__(self):
        print('-------INIT------')
        self.proxy_file_path = RusprofileSettings.objects.all()[0].proxy_file.path
        with open(self.proxy_file_path, 'r', encoding='utf-8') as proxy_file:
            proxies = [row.strip() for row in proxy_file.readlines()]
            self.proxies = sorted(proxies, key=lambda *args: random.random())

    def get_new_proxy(self):
        new_proxy = self.proxies.pop(0).split(':')

        ip, port, login, password = new_proxy[0], new_proxy[1], new_proxy[2], new_proxy[3]

        proxy_setting = {
            "http": f"http://{login}:{password}@{ip}:{port}",
            "https": f"https://{login}:{password}@{ip}:{port}"
        }

        if len(self.proxies) < 1:
            with open(self.proxy_file_path, 'r', encoding='utf-8') as proxy_file:
                proxies = [row.strip() for row in proxy_file.readlines()]
                self.proxies = sorted(proxies, key=lambda *args: random.random())

        return proxy_setting


def main(parser):

    proxy = Proxy()

    all_data = PageData.objects.filter(search_name__search_name__iexact=parser.search_file)

    with concurrent.futures.ThreadPoolExecutor(
            max_workers=RusprofileSettings.objects.all()[0].pool_workers
    ) as executor:
        executor.map(filter_and_search, [[i, parser, proxy] for i in all_data])

    parser.change_status(True)
